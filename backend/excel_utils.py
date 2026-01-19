"""
Excel Utilities Module

Handles reading and writing Excel files with specific sheet structures.
"""

from typing import List, Dict
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import io


class ExcelProcessor:
    """
    Processes Excel files according to application requirements.
    Reads ma_ho values from input sheet and creates RESULT sheet with findings.
    """
    
    def __init__(self):
        """Initialize the Excel processor."""
        self.workbook = None
        self.ma_ho_values = []
    
    def read_input_excel(self, file_content: bytes) -> List[str]:
        """
        Read Excel file and extract ma_ho values from column A.
        
        Args:
            file_content: Excel file content as bytes
            
        Returns:
            List of ma_ho values (excluding header)
            
        Raises:
            ValueError: If Excel file is invalid or missing required column
        """
        try:
            # Load workbook from bytes
            file_stream = io.BytesIO(file_content)
            self.workbook = load_workbook(file_stream)
            
            # Get the first (and should be only) sheet
            if not self.workbook.sheetnames:
                raise ValueError("Excel file has no sheets")
            
            sheet = self.workbook.active
            
            # Check if column A header is "ma_ho"
            header_cell = sheet.cell(row=1, column=1).value
            if not header_cell or str(header_cell).strip().lower() != "ma_ho":
                raise ValueError(
                    f"Expected column A header to be 'ma_ho', got '{header_cell}'"
                )
            
            # Extract all ma_ho values (skip header row)
            ma_ho_values = []
            for row in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value:  # Only include non-empty values
                    ma_ho_values.append(str(cell_value).strip())
            
            self.ma_ho_values = ma_ho_values
            return ma_ho_values
            
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")
    
    def create_result_sheet(self, results: List[Dict]) -> None:
        """
        Create a new sheet named "RESULT" with processing results.
        
        Args:
            results: List of dictionaries with keys:
                - ma_ho: The search term
                - found: "YES" or "NO"
                - file_name: Name of matching file (or empty)
                - file_id: Google Drive file ID (or empty)
                - folder_path: Path to the file (or empty)
        """
        if not self.workbook:
            raise ValueError("Workbook not loaded. Call read_input_excel first.")
        
        # Create new sheet named RESULT
        if "RESULT" in self.workbook.sheetnames:
            # Remove existing RESULT sheet if it exists
            del self.workbook["RESULT"]
        
        result_sheet = self.workbook.create_sheet("RESULT")
        
        # Write headers
        headers = ["ma_ho", "found", "file_name", "file_id", "folder_path"]
        for col_idx, header in enumerate(headers, start=1):
            result_sheet.cell(row=1, column=col_idx, value=header)
        
        # Write results
        for row_idx, result in enumerate(results, start=2):
            result_sheet.cell(row=row_idx, column=1, value=result.get("ma_ho", ""))
            result_sheet.cell(row=row_idx, column=2, value=result.get("found", "NO"))
            result_sheet.cell(row=row_idx, column=3, value=result.get("file_name", ""))
            result_sheet.cell(row=row_idx, column=4, value=result.get("file_id", ""))
            result_sheet.cell(row=row_idx, column=5, value=result.get("folder_path", ""))
        
        # Auto-adjust column widths for better readability
        for col_idx in range(1, 6):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in result_sheet[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            result_sheet.column_dimensions[column_letter].width = adjusted_width
    
    def save_to_bytes(self) -> bytes:
        """
        Save the workbook to bytes for download.
        
        Returns:
            Excel file content as bytes
        """
        if not self.workbook:
            raise ValueError("Workbook not loaded")
        
        output_stream = io.BytesIO()
        self.workbook.save(output_stream)
        output_stream.seek(0)
        return output_stream.read()
    
    def get_ma_ho_count(self) -> int:
        """
        Get the count of ma_ho values read from input.
        
        Returns:
            Number of ma_ho values
        """
        return len(self.ma_ho_values)
