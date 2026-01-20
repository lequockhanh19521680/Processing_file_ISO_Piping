"""
File I/O utility module for PDF Search Tool.

This module handles reading keywords from Excel files and writing results
back to Excel with proper formatting.
"""

import os
from typing import List, Dict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..config import KEYWORD_COLUMN_NAMES, RESULT_COLUMNS
from .logger import get_logger

logger = get_logger(__name__)


def read_keywords_from_excel(excel_path: str) -> List[str]:
    """
    Read keywords (Mã hố) from an Excel file.
    
    Expects an Excel file with a column named 'ma_ho', 'Mã hố', or similar
    containing the keywords to search for (one per row).
    
    Args:
        excel_path: Path to the input Excel file
        
    Returns:
        List of keywords extracted from the Excel file
        
    Raises:
        FileNotFoundError: If the Excel file does not exist
        ValueError: If the required column is not found or file is empty
    """
    logger.info(f"Reading keywords from Excel: {excel_path}")
    
    if not os.path.exists(excel_path):
        error_msg = f"Excel file not found: {excel_path}"
        logger.error(error_msg)
        raise FileNotFoundError(error_msg)
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_path)
        
        # Look for the keyword column (case-insensitive)
        keyword_column = None
        for col in df.columns:
            if col.lower() in KEYWORD_COLUMN_NAMES:
                keyword_column = col
                break
        
        if keyword_column is None:
            error_msg = f"No keyword column found. Expected one of: {KEYWORD_COLUMN_NAMES}"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Extract keywords (remove NaN and empty values)
        keywords = df[keyword_column].dropna().astype(str).str.strip().tolist()
        keywords = [k for k in keywords if k]  # Remove empty strings
        
        if not keywords:
            error_msg = "No keywords found in Excel file"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.info(f"Successfully loaded {len(keywords)} keywords")
        return keywords
        
    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}", exc_info=True)
        raise


def save_results_to_excel(
    results: List[Dict],
    output_path: str,
    input_excel_path: str
) -> None:
    """
    Save search results to an Excel file with a new RESULT sheet.
    
    Creates a new Excel file with:
    1. Original sheet(s) from the input file (preserved)
    2. New "RESULT" sheet with search results including Match_Type column
    
    Args:
        results: List of result dictionaries
        output_path: Path where the output Excel file will be saved
        input_excel_path: Path to the original input Excel file
        
    Raises:
        ValueError: If no results are available to save
    """
    logger.info(f"Saving results to Excel: {output_path}")
    
    if not results:
        error_msg = "No results to save"
        logger.error(error_msg)
        raise ValueError(error_msg)
    
    try:
        # Load the original workbook to preserve existing sheets
        workbook = load_workbook(input_excel_path)
        
        # Create or replace RESULT sheet
        if "RESULT" in workbook.sheetnames:
            logger.info("Removing existing RESULT sheet")
            del workbook["RESULT"]
        
        result_sheet = workbook.create_sheet("RESULT")
        
        # Write headers
        headers = ["ma_ho", "found", "file_name", "file_path", "Match_Type"]
        for col_idx, header in enumerate(headers, start=1):
            result_sheet.cell(row=1, column=col_idx, value=header)
        
        # Write results
        for row_idx, result in enumerate(results, start=2):
            result_sheet.cell(row=row_idx, column=1, value=result['ma_ho'])
            result_sheet.cell(row=row_idx, column=2, value=result['found'])
            result_sheet.cell(row=row_idx, column=3, value=result['file_name'])
            result_sheet.cell(row=row_idx, column=4, value=result['file_path'])
            result_sheet.cell(row=row_idx, column=5, value=result['match_type'])
        
        # Auto-adjust column widths for better readability
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in result_sheet[column_letter]:
                try:
                    if row.value and len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            result_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(output_path)
        
        # Log statistics
        found_count = sum(1 for r in results if r['found'] == 'YES')
        not_found_count = len(results) - found_count
        
        logger.info(f"Results successfully saved to {output_path}")
        logger.info(f"Total keywords processed: {len(results)}")
        logger.info(f"Keywords found: {found_count}")
        logger.info(f"Keywords not found: {not_found_count}")
        
    except Exception as e:
        logger.error(f"Error saving results to Excel: {str(e)}", exc_info=True)
        raise
