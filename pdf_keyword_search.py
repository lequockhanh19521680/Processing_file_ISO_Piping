#!/usr/bin/env python3
"""
PDF Keyword Search Script

A production-ready Python script that searches for specific keywords (Manhole Codes - "Mã hố")
from an Excel file across a large directory of PDF files using concurrent processing.

Features:
    - Concurrent processing with 50 workers using ThreadPoolExecutor
    - Recursive PDF scanning in directories and subdirectories
    - Keyword frequency counting (distinguishes single vs. multiple occurrences)
    - Professional progress bar using tqdm
    - Comprehensive logging to debug.log file
    - Excel input/output with openpyxl

Author: Generated for ISO Piping Project
Version: 1.0.0
"""

import os
import sys
import logging
from pathlib import Path
from typing import List, Dict, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import pdfplumber
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('debug.log', mode='a', encoding='utf-8'),
        logging.StreamHandler(sys.stderr)  # Errors to stderr, tqdm uses stdout
    ]
)
logger = logging.getLogger(__name__)


class PDFKeywordSearcher:
    """
    Main class for searching keywords in PDF files with concurrent processing.
    
    This class handles reading keywords from Excel, scanning PDF directories,
    extracting text from PDFs, counting keyword occurrences, and generating
    an output Excel file with results.
    
    Attributes:
        max_workers (int): Number of concurrent workers for PDF processing (default: 50)
        keywords (List[str]): List of keywords to search for
        results (List[Dict]): Search results for each keyword
    """
    
    def __init__(self, max_workers: int = 50):
        """
        Initialize the PDF Keyword Searcher.
        
        Args:
            max_workers (int): Maximum number of concurrent workers for ThreadPoolExecutor.
                             Default is 50 for optimal performance with thousands of files.
        """
        self.max_workers = max_workers
        self.keywords: List[str] = []
        self.results: List[Dict] = []
        logger.info(f"PDFKeywordSearcher initialized with {max_workers} workers")
    
    def read_keywords_from_excel(self, excel_path: str) -> List[str]:
        """
        Read keywords (Mã hố) from an Excel file.
        
        Expects an Excel file with a column named 'ma_ho' or 'Mã hố' containing
        the keywords to search for (one per row).
        
        Args:
            excel_path (str): Path to the input Excel file
            
        Returns:
            List[str]: List of keywords extracted from the Excel file
            
        Raises:
            FileNotFoundError: If the Excel file does not exist
            ValueError: If the required column is not found or file is empty
        """
        logger.info(f"Reading keywords from Excel: {excel_path}")
        
        if not os.path.exists(excel_path):
            error_msg = f"Excel file not found: {excel_path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        try:
            # Read Excel file
            df = pd.read_excel(excel_path)
            
            # Look for the keyword column (case-insensitive)
            keyword_column = None
            for col in df.columns:
                if col.lower() in ['ma_ho', 'mã hố', 'ma ho', 'keyword', 'keywords']:
                    keyword_column = col
                    break
            
            if keyword_column is None:
                error_msg = "No keyword column found. Expected column: 'ma_ho' or 'Mã hố'"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            # Extract keywords (remove NaN and empty values)
            keywords = df[keyword_column].dropna().astype(str).str.strip().tolist()
            keywords = [k for k in keywords if k]  # Remove empty strings
            
            if not keywords:
                error_msg = "No keywords found in Excel file"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            self.keywords = keywords
            logger.info(f"Successfully loaded {len(keywords)} keywords")
            return keywords
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}", exc_info=True)
            raise
    
    def find_pdf_files(self, root_directory: str) -> List[str]:
        """
        Recursively find all PDF files in a directory and its subdirectories.
        
        Args:
            root_directory (str): Root directory to search for PDF files
            
        Returns:
            List[str]: List of absolute paths to PDF files
            
        Raises:
            FileNotFoundError: If the root directory does not exist
        """
        logger.info(f"Scanning for PDF files in: {root_directory}")
        
        if not os.path.exists(root_directory):
            error_msg = f"Directory not found: {root_directory}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        pdf_files = []
        root_path = Path(root_directory)
        
        # Recursively find all PDF files
        for pdf_file in root_path.rglob("*.pdf"):
            if pdf_file.is_file():
                pdf_files.append(str(pdf_file.absolute()))
        
        # Also check for .PDF (uppercase extension)
        for pdf_file in root_path.rglob("*.PDF"):
            if pdf_file.is_file() and str(pdf_file.absolute()) not in pdf_files:
                pdf_files.append(str(pdf_file.absolute()))
        
        logger.info(f"Found {len(pdf_files)} PDF files")
        return sorted(pdf_files)
    
    def extract_text_from_pdf(self, pdf_path: str) -> Optional[str]:
        """
        Extract text content from a PDF file using pdfplumber.
        
        Args:
            pdf_path (str): Path to the PDF file
            
        Returns:
            Optional[str]: Extracted text from all pages, or None if extraction fails
        """
        try:
            text_parts = []
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)
            
            full_text = " ".join(text_parts)
            logger.debug(f"Extracted {len(full_text)} characters from {os.path.basename(pdf_path)}")
            return full_text
            
        except Exception as e:
            logger.warning(f"Failed to extract text from {pdf_path}: {str(e)}")
            return None
    
    def count_keyword_occurrences(self, text: str, keyword: str) -> int:
        """
        Count how many times a keyword appears in the text (case-insensitive).
        
        Args:
            text (str): Text to search in
            keyword (str): Keyword to search for
            
        Returns:
            int: Number of times the keyword appears in the text
        """
        if not text or not keyword:
            return 0
        
        # Case-insensitive search
        text_lower = text.lower()
        keyword_lower = keyword.lower()
        
        count = text_lower.count(keyword_lower)
        return count
    
    def search_keyword_in_pdf(self, pdf_path: str, keyword: str) -> Dict:
        """
        Search for a keyword in a single PDF file and count occurrences.
        
        This function is designed to be called by multiple workers concurrently.
        
        Args:
            pdf_path (str): Path to the PDF file
            keyword (str): Keyword to search for
            
        Returns:
            Dict: Dictionary containing:
                - pdf_path (str): Path to the PDF file
                - keyword (str): The keyword searched
                - count (int): Number of occurrences found
                - found (bool): Whether the keyword was found
        """
        try:
            text = self.extract_text_from_pdf(pdf_path)
            if text is None:
                return {
                    'pdf_path': pdf_path,
                    'keyword': keyword,
                    'count': 0,
                    'found': False
                }
            
            count = self.count_keyword_occurrences(text, keyword)
            return {
                'pdf_path': pdf_path,
                'keyword': keyword,
                'count': count,
                'found': count > 0
            }
            
        except Exception as e:
            logger.error(f"Error searching keyword '{keyword}' in {pdf_path}: {str(e)}", exc_info=True)
            return {
                'pdf_path': pdf_path,
                'keyword': keyword,
                'count': 0,
                'found': False
            }
    
    def process_keywords(self, pdf_files: List[str]) -> List[Dict]:
        """
        Process all keywords against all PDF files using concurrent execution.
        
        Uses ThreadPoolExecutor with the configured number of workers to process
        multiple PDFs in parallel. Displays a professional progress bar using tqdm.
        
        Args:
            pdf_files (List[str]): List of PDF file paths to search
            
        Returns:
            List[Dict]: Results for each keyword with match information
        """
        logger.info(f"Starting concurrent processing with {self.max_workers} workers")
        logger.info(f"Processing {len(self.keywords)} keywords across {len(pdf_files)} PDF files")
        
        results = []
        
        # Create a tqdm progress bar
        with tqdm(total=len(self.keywords), 
                  desc="Processing Keywords", 
                  unit="keyword",
                  ncols=100) as pbar:
            
            for keyword_idx, keyword in enumerate(self.keywords, 1):
                logger.info(f"Processing keyword {keyword_idx}/{len(self.keywords)}: {keyword}")
                
                # Update tqdm description with current keyword
                pbar.set_description(f"Processing [{keyword_idx}/{len(self.keywords)}] | Code: {keyword[:20]}")
                
                best_match = None
                max_count = 0
                
                # Process all PDFs for this keyword concurrently
                with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                    # Submit all PDF search tasks for this keyword
                    future_to_pdf = {
                        executor.submit(self.search_keyword_in_pdf, pdf_path, keyword): pdf_path
                        for pdf_path in pdf_files
                    }
                    
                    # Process completed tasks
                    for future in as_completed(future_to_pdf):
                        pdf_path = future_to_pdf[future]
                        try:
                            result = future.result()
                            
                            # Keep track of the best match (highest count)
                            if result['found'] and result['count'] > max_count:
                                max_count = result['count']
                                best_match = result
                                
                        except Exception as e:
                            logger.error(f"Task failed for {pdf_path}: {str(e)}", exc_info=True)
                
                # Prepare result for this keyword
                if best_match:
                    # Determine Match_Count/Status based on frequency
                    # max_count is guaranteed to be > 0 here since best_match exists
                    if max_count == 1:
                        status = "1"
                    else:  # max_count >= 2
                        status = f"{max_count}"  # Show actual count for 2+
                    
                    results.append({
                        'ma_ho': keyword,
                        'found': 'YES',
                        'file_name': os.path.basename(best_match['pdf_path']),
                        'file_path': best_match['pdf_path'],
                        'match_count': max_count,
                        'status': status
                    })
                    logger.info(f"✓ Keyword '{keyword}' found {max_count} time(s) in {os.path.basename(best_match['pdf_path'])}")
                else:
                    results.append({
                        'ma_ho': keyword,
                        'found': 'NO',
                        'file_name': '',
                        'file_path': '',
                        'match_count': 0,
                        'status': '0'
                    })
                    logger.info(f"✗ Keyword '{keyword}' not found")
                
                # Update progress bar
                pbar.update(1)
        
        self.results = results
        logger.info(f"Processing complete. Found matches for {sum(1 for r in results if r['found'] == 'YES')} keywords")
        return results
    
    def save_results_to_excel(self, output_path: str, input_excel_path: str) -> None:
        """
        Save search results to an Excel file with a new RESULT sheet.
        
        Creates a new Excel file with:
        1. Original sheet(s) from the input file (preserved)
        2. New "RESULT" sheet with search results including Match_Count column
        
        Args:
            output_path (str): Path where the output Excel file will be saved
            input_excel_path (str): Path to the original input Excel file
            
        Raises:
            ValueError: If no results are available to save
        """
        logger.info(f"Saving results to Excel: {output_path}")
        
        if not self.results:
            error_msg = "No results to save. Run process_keywords first."
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        try:
            # Load the original workbook to preserve existing sheets
            workbook = load_workbook(input_excel_path)
            
            # Create or replace RESULT sheet
            if "RESULT" in workbook.sheetnames:
                logger.info("Removing existing RESULT sheet")
                del workbook["RESULT"]
            
            result_sheet = workbook.create_sheet("RESULT")
            
            # Write headers
            headers = ["ma_ho", "found", "file_name", "file_path", "match_count", "status"]
            for col_idx, header in enumerate(headers, start=1):
                result_sheet.cell(row=1, column=col_idx, value=header)
            
            # Write results
            for row_idx, result in enumerate(self.results, start=2):
                result_sheet.cell(row=row_idx, column=1, value=result['ma_ho'])
                result_sheet.cell(row=row_idx, column=2, value=result['found'])
                result_sheet.cell(row=row_idx, column=3, value=result['file_name'])
                result_sheet.cell(row=row_idx, column=4, value=result['file_path'])
                result_sheet.cell(row=row_idx, column=5, value=result['match_count'])
                result_sheet.cell(row=row_idx, column=6, value=result['status'])
            
            # Auto-adjust column widths for better readability
            for col_idx in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for row in result_sheet[column_letter]:
                    try:
                        if row.value and len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                result_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            workbook.save(output_path)
            logger.info(f"Results successfully saved to {output_path}")
            logger.info(f"Total keywords processed: {len(self.results)}")
            logger.info(f"Keywords found: {sum(1 for r in self.results if r['found'] == 'YES')}")
            logger.info(f"Keywords not found: {sum(1 for r in self.results if r['found'] == 'NO')}")
            
        except Exception as e:
            logger.error(f"Error saving results to Excel: {str(e)}", exc_info=True)
            raise


def main():
    """
    Main entry point for the PDF Keyword Search script.
    
    Usage:
        python pdf_keyword_search.py <input_excel> <pdf_directory> <output_excel>
    
    Arguments:
        input_excel: Path to Excel file containing keywords (ma_ho column)
        pdf_directory: Root directory containing PDF files to search
        output_excel: Path where the output Excel file will be saved
    
    Example:
        python pdf_keyword_search.py input.xlsx ./pdf_files/ output.xlsx
    """
    print("=" * 80)
    print("PDF Keyword Search - Production Script")
    print("Version 1.0.0")
    print("=" * 80)
    print()
    
    # Check command line arguments
    if len(sys.argv) != 4:
        print("Usage: python pdf_keyword_search.py <input_excel> <pdf_directory> <output_excel>")
        print()
        print("Arguments:")
        print("  input_excel    : Path to Excel file with keywords (ma_ho column)")
        print("  pdf_directory  : Root directory containing PDF files")
        print("  output_excel   : Path for output Excel file")
        print()
        print("Example:")
        print("  python pdf_keyword_search.py input.xlsx ./pdf_files/ output.xlsx")
        sys.exit(1)
    
    input_excel = sys.argv[1]
    pdf_directory = sys.argv[2]
    output_excel = sys.argv[3]
    
    logger.info("=" * 80)
    logger.info("Starting PDF Keyword Search")
    logger.info(f"Input Excel: {input_excel}")
    logger.info(f"PDF Directory: {pdf_directory}")
    logger.info(f"Output Excel: {output_excel}")
    logger.info("=" * 80)
    
    try:
        # Initialize searcher with 50 workers
        print(f"[1/5] Initializing with 50 concurrent workers...")
        searcher = PDFKeywordSearcher(max_workers=50)
        
        # Read keywords from Excel
        print(f"[2/5] Reading keywords from Excel file...")
        keywords = searcher.read_keywords_from_excel(input_excel)
        print(f"      Loaded {len(keywords)} keywords")
        
        # Find PDF files
        print(f"[3/5] Scanning for PDF files in directory...")
        pdf_files = searcher.find_pdf_files(pdf_directory)
        print(f"      Found {len(pdf_files)} PDF files")
        
        if not pdf_files:
            print("      WARNING: No PDF files found!")
            logger.warning("No PDF files found in directory")
            # Create empty results
            searcher.results = [
                {
                    'ma_ho': kw,
                    'found': 'NO',
                    'file_name': '',
                    'file_path': '',
                    'match_count': 0,
                    'status': '0'
                }
                for kw in keywords
            ]
        else:
            # Process keywords
            print(f"[4/5] Processing keywords (this may take a while)...")
            results = searcher.process_keywords(pdf_files)
        
        # Save results
        print(f"[5/5] Saving results to Excel file...")
        searcher.save_results_to_excel(output_excel, input_excel)
        
        print()
        print("=" * 80)
        print("✓ Processing Complete!")
        print("=" * 80)
        print(f"Results saved to: {output_excel}")
        print(f"Total keywords: {len(searcher.results)}")
        print(f"Keywords found: {sum(1 for r in searcher.results if r['found'] == 'YES')}")
        print(f"Keywords not found: {sum(1 for r in searcher.results if r['found'] == 'NO')}")
        print(f"Log file: debug.log")
        print("=" * 80)
        
        logger.info("Script completed successfully")
        
    except Exception as e:
        print()
        print("=" * 80)
        print("✗ ERROR OCCURRED")
        print("=" * 80)
        print(f"Error: {str(e)}")
        print(f"Check debug.log for detailed information")
        print("=" * 80)
        logger.error(f"Script failed: {str(e)}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
