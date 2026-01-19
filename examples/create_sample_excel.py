"""
Create a sample Excel file for testing the ISO Piping File Processor.
This script creates an example input file with ma_ho values.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_sample_excel(filename="sample_input.xlsx"):
    """Create a sample Excel file with ma_ho values."""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Style for header
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add header
    ws['A1'] = 'ma_ho'
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=12)
    ws['A1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    
    # Sample ma_ho values
    sample_values = [
        "ISO-P-001",
        "ISO-P-002",
        "ISO-P-003",
        "VALVE-A001",
        "VALVE-A002",
        "PIPE-2024-001",
        "PIPE-2024-002",
        "FLANGE-B123",
        "FITTING-C456",
        "SUPPORT-D789"
    ]
    
    # Add sample values
    for idx, value in enumerate(sample_values, start=2):
        ws[f'A{idx}'] = value
    
    # Adjust column width
    ws.column_dimensions['A'].width = 20
    
    # Save workbook
    wb.save(filename)
    print(f"✓ Sample Excel file created: {filename}")
    print(f"  Contains {len(sample_values)} ma_ho values")
    print(f"  You can use this file to test the application")

if __name__ == "__main__":
    create_sample_excel()
