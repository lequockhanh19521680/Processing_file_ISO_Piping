"""
Excel processing utilities for reading and writing hole codes.
"""
import pandas as pd
import openpyxl
from typing import List, Dict


def read_excel_codes(file_path: str) -> List[str]:
    """
    Read ma_ho codes from the Excel file.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        List of ma_ho codes
    """
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        
        if 'ma_ho' not in df.columns:
            raise ValueError("Column 'ma_ho' not found in Excel file")
        
        # Get non-null ma_ho values and convert to string
        codes = df['ma_ho'].dropna().astype(str).tolist()
        
        return codes
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")


def write_result_sheet(file_path: str, results: List[Dict[str, str]]) -> str:
    """
    Write results to a new sheet named RESULT in the Excel file.
    
    Args:
        file_path: Path to the Excel file
        results: List of dictionaries containing results for each ma_ho
        
    Returns:
        Path to the updated Excel file
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        
        # Remove RESULT sheet if it already exists
        if 'RESULT' in workbook.sheetnames:
            del workbook['RESULT']
        
        # Create new RESULT sheet
        result_sheet = workbook.create_sheet('RESULT')
        
        # Write headers
        headers = ['ma_ho', 'found', 'file_name', 'file_id', 'folder_path']
        result_sheet.append(headers)
        
        # Write results
        for result in results:
            row = [
                result.get('ma_ho', ''),
                result.get('found', 'NO'),
                result.get('file_name', ''),
                result.get('file_id', ''),
                result.get('folder_path', '')
            ]
            result_sheet.append(row)
        
        # Save the workbook
        workbook.save(file_path)
        
        return file_path
    except Exception as e:
        raise Exception(f"Error writing result sheet: {str(e)}")
